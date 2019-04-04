# START by running: pip install openpyxl

from pprint import pprint as pp
import sys
import os
import openpyxl as xl

# FUNCTIONS #

def check_exists(filename):
    if not os.path.exists(filename):
        print(f"ERROR: {filename} doesn't exist")
        exit()

def find_col_with_val(wksheet, val, row_number):
    '''
    Given a value and a row_number, return the column
    in which the FIRST occurrence of value is found.
    '''
    row = wksheet[row_number]
    for cell in row:
        #print(cell.value)
        if cell.value.lower() == val.lower():
            return cell.column_letter
    return ''

def find_row_with_val(wksheet, val, col_letter):
    '''
    Given a value and a col_letter, return the row
    in which the FIRST occurrence of value is found.
    '''
    col = wksheet[col_letter]
    for cell in col:
        #print(cell.value)
        if cell.value.lower() == val.lower():
            return cell.row
    return -1

# python gradeit.py [newgrades] [gradesheet] [translation]

#pp(sys.argv)

if len(sys.argv) != 4:
    print('USAGE: python gradeit.py [newgrades] [gradesheet] [translation]')
    exit()

newgrades = sys.argv[1]
gradesheet = sys.argv[2]
translation = sys.argv[3]

#print(newgrades, gradesheet, translation)

check_exists(newgrades)
check_exists(gradesheet)
check_exists(translation)

'''
1) open newgrades
2) for each grade
    2a) read id and grade
    2b) find email associated with id (using translation)
    2c) using gradesheet find right col and row for assignment and email
    2d) write grade in correct cell
3) save gradesheet
'''

ng_book = xl.load_workbook(newgrades)
ng_sheet = ng_book.worksheets[0]

# Read grades from newgrades
#for row in ng_sheet.rows:
for row in ng_sheet.iter_rows(2):
    #pp(row)
    ghid = row[0].value
    grade = row[1].value
    print(ghid, grade)

    # FIND email for the id
    tr_book = xl.load_workbook(translation)
    tr_sheet = tr_book.worksheets[0]
    email = 'NA'
    for tr_row in tr_sheet.iter_rows(2):
        if ghid == tr_row[1].value:
            email = tr_row[0].value
            break
    tr_book.close()

    if email == 'NA':
        print(f"WARNING: {ghid} not found")
        continue

    print(f"{ghid} actually is {email}")

    part_a = newgrades.split(os.sep) # ["..","..","HW4.xlsx"]
    fn = part_a[-1] # "HW4.xlsx"
    part_b = fn.split('.') # ["HW4", "xlsx"]
    assn = part_b[0]

    print(f'PROCESSING: {assn}')

    # DONE FOR TODAY

    # FIND email for the id
    gr_book = xl.load_workbook(gradesheet)
    gr_sheet = gr_book.worksheets[0]
    
    col = find_col_with_val(gr_sheet, assn, 1)
    row = find_row_with_val(gr_sheet, email, 'B')
    coord = col + str(row)
    print(f"Changing {coord} to {grade}")

    gr_sheet[coord].value = grade
    
    gr_book.save(gradesheet)
    gr_book.close()

ng_book.close()
