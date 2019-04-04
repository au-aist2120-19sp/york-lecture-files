import os
import sys
import openpyxl as xl
from pprint import pprint as pp

# FUNCTIONS #

def check_exists(filename):
    if not os.path.exists(filename):
        print(f'ERROR: {filename} does not exist')
        exit()

def get_col_for_val(wksheet, val, row_num):
    '''
    Find the first value in a worksheet row and return
    the associated column letter.
    '''
    for cell in wksheet[row_num]:
        # (CELL1, CELL2, CELL3,...)
        if cell.value.lower() == val.lower():
            return cell.column_letter
    return ''

def get_row_for_val(wksheet, val, col_letter):
    '''
    Find the first value in a worksheet column and return
    the associated row number.
    '''
    for cell in wksheet[col_letter]:
        # (CELL1, CELL2, CELL3,...)
        if cell.value.lower() == val.lower():
            return cell.row
    return -1

# python gradeit.py [newgrades] [gradesheet] [translation]
# pp(sys.argv)

if len(sys.argv) != 4:
    print('USAGE: python gradeit.py [newgrades] [gradesheet] [translation]')
    exit()

newgrades = sys.argv[1]
gradesheet = sys.argv[2]
translation = sys.argv[3]

check_exists(newgrades)
check_exists(gradesheet)
check_exists(translation)

# open the newgrades file
# for every new grade
#    find the translation (i.e., email address) for the id
#    update the gradesheet for the student

ng_book = xl.load_workbook(newgrades)
#ng_sheet = ng_book.active
ng_sheet = ng_book.worksheets[0]

#for ng_row in ng_sheet.rows:
for ng_row in ng_sheet.iter_rows(2):
    #print(ng_row)
    ghid = ng_row[0].value
    grade = ng_row[1].value
    print(ghid, grade)

    # FIND associated email
    tr_book = xl.load_workbook(translation)
    tr_sheet = tr_book.worksheets[0]
    email = 'NA'
    for tr_row in tr_sheet.iter_rows(2):
        if tr_row[1].value == ghid:
            email = tr_row[0].value
            break
    print(f'{ghid}==>{email}')
    if email == 'NA':
        print(f"WARNING: {ghid} not found!")
        continue

    tr_book.close()

    #assn = 'HW4'
    #print(newgrades)
    #..\..\HW4.xlsx
    pt_a = newgrades.split(os.sep)
    #["..", "..", "HW4.xlsx"]
    fn = pt_a[-1]
    #"HW4.xlsx"
    pt_b = fn.split('.')
    #["HW4", "xlsx"] 
    assn = pt_b[0]
    #"HW4"

    print(f"Now processing grade for {assn}")

    # STOPPING UNTIL THURSDAY

    gr_book = xl.load_workbook(gradesheet)
    gr_sheet = gr_book.worksheets[0]

    col = get_col_for_val(gr_sheet, assn, '1')
    #print(col)
    row = get_row_for_val(gr_sheet, email, 'B')
    #print(row)

    coord = col + str(row)

    print(f"Setting {coord} to {grade}")
    cell = gr_sheet[coord]
    cell.value = grade

    gr_book.save(gradesheet)
    gr_book.close()

ng_book.close()
